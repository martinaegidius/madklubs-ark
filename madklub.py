#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Sep 25 15:33:47 2022

@author: max
"""

from openpyxl import Workbook
import datetime
import calendar
from dateutil import relativedelta
import os 
import pickle


f = os.path.dirname(__file__)+"/madklub.xlsx"
if os.path.isfile(f):
    print("File existed already. Cleaning up")
    os.rename(f,f+".old")
    
filename = os.path.dirname(__file__)+"/madklub.xlsx"
 

# get current datetime
dt = datetime.datetime.now()
print('Datetime is:', dt)
nextMonthDate = dt + relativedelta.relativedelta(months=1, day=1)#get first date of next month 

#today = datetime.date.today()
daysInMonth= calendar.monthrange(nextMonthDate.year, nextMonthDate.month)[1] #get days in month of next month
wb = Workbook()
ws = wb.active

weekdays = {0:"Mon",1:"Tue",2:"Wed",3:"Thu",4:"Fri",5:"Sat",6:"Sun"}
next_month = (dt.replace(day=1) + datetime.timedelta(days=32*1)).replace(day=1)
print("Next month object: ",next_month)

#print("Next month object: ",nextMonthDate)
day_startidx = next_month.weekday()
#day_startidx = nextMonthDate.weekday()

dayidx = day_startidx
food_days_in_month = 0
for day in range(1,daysInMonth+1):
    if(dayidx>6):
        dayidx = 0
        
    if(dayidx!=4 and dayidx !=5):
        food_days_in_month += 1
    rowno = 3+day-1
    if day<10:
        date = "0"+str(day)
    else: 
        date = str(day)
    tmp_stringA = 'A'+str(rowno)
    tmp_stringB = 'B'+str(rowno)
    ws[tmp_stringA] = date
    ws[tmp_stringB] = weekdays[dayidx]
    print(weekdays[dayidx])
    dayidx += 1

print("Number of days with foodclub next month: ",food_days_in_month)

if(food_days_in_month < 0):
    print("Someone will not get a foodclub in this month. Thus, you must push to zero position of list the room which gets no foodclub before running next time")

if(food_days_in_month > 0):
    with open("/home/max/Documents/Madklub_ark/roomlist","rb") as fp: 
        roomlist = pickle.load(fp)
    
    with open(r'/home/max/Documents/Madklub_ark/holy_roomlist/rooms_before.txt', 'w') as fp:
        for item in roomlist:
            # write each item on a new line
            fp.write("%s\n" % item)
    
    num_doubles_next_month = food_days_in_month-len(roomlist)
    
    print("Number double foodclubs next month: ",num_doubles_next_month)
    ws['E56'] = "Next month following rooms have two foodclubs"
    column_iter = ["F","G","H","I","J"]
    for i in range(num_doubles_next_month):
        pos = column_iter[i]+str(58)
        room = roomlist.pop(0)
        roomlist.append(room)
        ws[pos] = room
    
    with open("/home/max/Documents/Madklub_ark/roomlist","wb") as fp: #pickle again
        pickle.dump(roomlist,fp)
        
    with open(r'/home/max/Documents/Madklub_ark/holy_roomlist/rooms_after.txt', 'w') as fp:
        for item in roomlist:
            # write each item on a new line
            fp.write("%s\n" % item)
    
        
wb.save(filename)#save new sheet
print("Generation complete. Check file")